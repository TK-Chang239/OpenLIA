"""WorkbookTemplate — structured multi-sheet xlsx workbook generator.

Encapsulates institutional financial-modeling conventions: sheet layout, number
formats, bold totals, frozen panes, conditional formatting, and chart embedding.

Two pre-built templates ship here (skeleton implementations):
  - EquityResearchInitiation: full initiation workbook (Cover, Assumptions, DCF,
    Sensitivity, Scenarios, Comparables, SOTP, Cost of Capital, Decision, Citations)
  - EarningsUpdate: abbreviated update workbook (Cover, Financials, EPS Walk, Guidance)

Full template implementations land in Phase 3; this PR ships the class API and
infrastructure so workbook_builder can depend on it immediately.

Design ref: planning/2026-05-21-equity-research-helpers-design.md §2.5
"""

from __future__ import annotations

import io
from datetime import date
from typing import Any, ClassVar, Literal

SheetType = Literal["table", "chart", "narrative", "cover", "assumptions"]

_NUMBER_FORMATS: dict[str, str] = {
    "currency": "#,##0.00",
    "currency_millions": "#,##0.0,,",
    "pct": "0.0%",
    "pct2": "0.00%",
    "integer": "#,##0",
    "multiple": "0.0x",
    "ratio": "0.00",
}


class SheetSpec:
    """Configuration for a single workbook sheet."""

    def __init__(
        self,
        name: str,
        sheet_type: SheetType,
        layout_config: dict[str, Any] | None = None,
    ) -> None:
        self.name = name
        self.sheet_type: SheetType = sheet_type
        self.layout_config: dict[str, Any] = layout_config or {}
        # Rows written by add_data(), render()
        self._rows: list[list[Any]] = []
        self._headers: list[str] = []

    def set_headers(self, headers: list[str]) -> None:
        self._headers = headers

    def add_row(self, row: list[Any]) -> None:
        self._rows.append(row)

    @property
    def row_count(self) -> int:
        return len(self._rows)

    @property
    def col_count(self) -> int:
        if self._headers:
            return len(self._headers)
        if self._rows:
            return max(len(r) for r in self._rows)
        return 0

    @property
    def cell_count(self) -> int:
        if not self._rows:
            return 0
        header_row = 1 if self._headers else 0
        return (len(self._rows) + header_row) * self.col_count


class WorkbookTemplate:
    """Structured multi-sheet xlsx workbook with institutional formatting.

    Usage::

        wb = WorkbookTemplate(
            company_name="Apple",
            ticker="AAPL",
            currency="USD",
            report_date="2026-05-21",
        )
        wb.add_sheet("DCF", sheet_type="table", layout_config={"freeze_panes": "B2"})
        # populate sheets via embed_artifact() or direct row writes
        xlsx_bytes = wb.to_bytes()

    Full API for workbook_builder helper:

        wb.add_sheet(name, sheet_type, layout_config)  # declare sheet
        wb.write_headers(sheet_name, headers)           # write header row
        wb.write_row(sheet_name, row)                   # append a data row
        wb.embed_artifact(sheet_name, artifact)         # auto-populate from helper output
        wb.validate()                                   # check constraints
        bytes_ = wb.to_bytes()                          # render to xlsx bytes
        wb.save(path)                                   # write to file
        info = wb.metadata()                            # sheet/cell/size summary
    """

    def __init__(
        self,
        company_name: str,
        ticker: str,
        currency: str = "USD",
        report_date: str | date | None = None,
    ) -> None:
        self.company_name = company_name
        self.ticker = ticker
        self.currency = currency
        self.report_date: str = str(report_date) if report_date else date.today().isoformat()
        self._sheets: dict[str, SheetSpec] = {}
        self._sheet_order: list[str] = []

    # ------------------------------------------------------------------
    # Sheet management
    # ------------------------------------------------------------------

    def add_sheet(
        self,
        name: str,
        sheet_type: SheetType = "table",
        layout_config: dict[str, Any] | None = None,
    ) -> SheetSpec:
        """Declare a new sheet. Returns the SheetSpec for further configuration.

        Args:
            name: Sheet tab name. Max 31 chars (Excel limit).
            sheet_type: "table" | "chart" | "narrative" | "cover" | "assumptions".
            layout_config: Optional extra hints (freeze_panes, column_widths, etc.).

        Raises:
            ValueError: If sheet with same name already declared.
        """
        if len(name) > 31:
            raise ValueError(f"Sheet name {name!r} exceeds 31-char Excel limit")
        if name in self._sheets:
            raise ValueError(f"Sheet {name!r} already declared")
        spec = SheetSpec(name=name, sheet_type=sheet_type, layout_config=layout_config or {})
        self._sheets[name] = spec
        self._sheet_order.append(name)
        return spec

    def _get_or_add_sheet(self, name: str, sheet_type: SheetType = "table") -> SheetSpec:
        if name not in self._sheets:
            return self.add_sheet(name, sheet_type=sheet_type)
        return self._sheets[name]

    def write_headers(self, sheet_name: str, headers: list[str]) -> None:
        """Set column headers for a sheet."""
        self._sheets[sheet_name].set_headers(headers)

    def write_row(self, sheet_name: str, row: list[Any]) -> None:
        """Append a data row to a sheet."""
        self._sheets[sheet_name].add_row(row)

    # ------------------------------------------------------------------
    # Artifact embedding — auto-populated from helper output dicts
    # ------------------------------------------------------------------

    def embed_artifact(self, sheet_name: str, artifact: dict[str, Any]) -> None:
        """Populate a sheet from a helper output artifact.

        Inspects artifact for recognizable shapes:
        - rows / headers (table-style) — written directly
        - markdown_table — parsed into rows
        - summary dict — flattened as key/value pairs

        sheet_name is created if it does not already exist.
        """
        spec = self._get_or_add_sheet(sheet_name, sheet_type="table")

        # Table-style artifact (headers + rows)
        if "headers" in artifact and "rows" in artifact:
            spec.set_headers(list(artifact["headers"]))
            for row in artifact.get("rows", []):
                spec.add_row(list(row))
            return

        # markdown_table — parse pipe-delimited markdown
        if "markdown_table" in artifact:
            _parse_markdown_table_into(spec, artifact["markdown_table"])
            return

        # summary dict — flatten as key/value pairs
        if "summary" in artifact and isinstance(artifact["summary"], dict):
            spec.set_headers(["Metric", "Value"])
            for k, v in artifact["summary"].items():
                spec.add_row([k, v])
            return

        # Fallback: flatten the top-level dict
        spec.set_headers(["Key", "Value"])
        for k, v in artifact.items():
            if not isinstance(v, (dict, list)):
                spec.add_row([k, v])

    # ------------------------------------------------------------------
    # High-level helpers called by workbook_builder
    # ------------------------------------------------------------------

    def add_dcf(self, dcf_artifact: dict[str, Any]) -> None:
        """Write DCF schedule to a 'DCF' sheet."""
        spec = self._get_or_add_sheet("DCF", sheet_type="table")
        if dcf_artifact.get("fcff_schedule"):
            schedule = dcf_artifact["fcff_schedule"]
            if schedule:
                first = schedule[0]
                spec.set_headers(list(first.keys()))
                for row_dict in schedule:
                    spec.add_row(list(row_dict.values()))
        else:
            self.embed_artifact("DCF", dcf_artifact)

    def add_comparables(self, comparables_artifact: dict[str, Any]) -> None:
        """Write comparables peer table to a 'Comparables' sheet."""
        self.embed_artifact("Comparables", comparables_artifact)

    def add_sensitivity(self, sensitivity_artifact: dict[str, Any]) -> None:
        """Write 2-D sensitivity grid to a 'Sensitivity' sheet."""
        self.embed_artifact("Sensitivity", sensitivity_artifact)

    def add_scenarios(self, scenarios_artifact: dict[str, Any]) -> None:
        """Write scenario analysis to a 'Scenarios' sheet."""
        self.embed_artifact("Scenarios", scenarios_artifact)

    def add_sotp(self, sotp_artifact: dict[str, Any]) -> None:
        """Write SOTP segment table to an 'SOTP' sheet."""
        self.embed_artifact("SOTP", sotp_artifact)

    def add_decision(self, decision_artifact: dict[str, Any]) -> None:
        """Write decision layer output (rating, PT, ETR) to a 'Decision' sheet."""
        self.embed_artifact("Decision", decision_artifact)

    def add_panel(self, panel_artifact: dict[str, Any], sheet_name: str | None = None) -> None:
        """Write an arbitrary panel artifact to a sheet.

        sheet_name defaults to artifact.get('artifact_type') or 'Panel'.
        """
        name = sheet_name or panel_artifact.get("artifact_type", "Panel")
        # Truncate to 31 chars
        name = str(name)[:31]
        self.embed_artifact(name, panel_artifact)

    # ------------------------------------------------------------------
    # Render / serialise
    # ------------------------------------------------------------------

    def render(self, *, write_cover: bool = True) -> Any:
        """Produce an openpyxl Workbook from the declared sheets.

        Args:
            write_cover: If True and no 'Cover' sheet declared, prepend a cover sheet.

        Returns:
            openpyxl.Workbook (not yet saved to disk).
        """
        import openpyxl  # type: ignore[import-untyped]
        from openpyxl.styles import Font, PatternFill  # type: ignore[import-untyped]

        wb = openpyxl.Workbook()
        # Remove default empty sheet
        default = wb.active
        if default is not None:
            wb.remove(default)

        # Prepend cover sheet if requested and not already declared
        if write_cover and "Cover" not in self._sheets:
            self._write_cover_sheet(wb)

        for sheet_name in self._sheet_order:
            spec = self._sheets[sheet_name]
            ws = wb.create_sheet(title=sheet_name)
            self._render_sheet(ws, spec, Font=Font, PatternFill=PatternFill)

        # Write declared cover last (if it was in the sheet order) or first (if injected)
        return wb

    def _write_cover_sheet(self, wb: Any) -> None:
        """Prepend a cover sheet with company/ticker/date/currency."""
        from openpyxl.styles import Font  # type: ignore[import-untyped]

        ws = wb.create_sheet(title="Cover", index=0)
        ws["A1"] = self.company_name
        ws["A1"].font = Font(bold=True, size=16)
        ws["A2"] = f"Ticker: {self.ticker}"
        ws["A3"] = f"Currency: {self.currency}"
        ws["A4"] = f"Report Date: {self.report_date}"

        # TOC
        ws["A6"] = "Sheet"
        ws["B6"] = "Contents"
        ws["A6"].font = Font(bold=True)
        ws["B6"].font = Font(bold=True)
        for i, name in enumerate(self._sheet_order, start=7):
            spec = self._sheets[name]
            ws[f"A{i}"] = name
            ws[f"B{i}"] = f"{spec.sheet_type} — {spec.row_count} rows"
        ws.freeze_panes = "A2"

    def _render_sheet(self, ws: Any, spec: SheetSpec, Font: Any, PatternFill: Any) -> None:
        """Write headers + rows from a SheetSpec onto an openpyxl worksheet."""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        row_offset = 1
        if spec._headers:
            for col_idx, h in enumerate(spec._headers, start=1):
                cell = ws.cell(row=row_offset, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
            row_offset += 1

        for row in spec._rows:
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=row_offset, column=col_idx, value=val)
            row_offset += 1

        # Freeze panes per layout_config or default to A2 if headers present
        freeze = spec.layout_config.get("freeze_panes")
        if freeze:
            ws.freeze_panes = freeze
        elif spec._headers:
            ws.freeze_panes = "A2"

        # Column width hints
        col_widths: dict[str, int] = spec.layout_config.get("column_widths", {})
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

    def validate(self) -> list[str]:
        """Check declared sheets for basic constraints. Return list of warnings.

        Does not raise — returns warnings so caller can decide.
        """
        warnings: list[str] = []
        if not self._sheets:
            warnings.append("No sheets declared; workbook will be empty.")
        for name, spec in self._sheets.items():
            if len(name) > 31:
                warnings.append(f"Sheet name {name!r} exceeds 31-char Excel limit.")
            if spec.sheet_type == "table" and not spec._headers and not spec._rows:
                warnings.append(f"Sheet {name!r} is declared as table but has no data.")
        return warnings

    def to_bytes(self) -> bytes:
        """Render workbook to raw xlsx bytes."""
        wb = self.render()
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def save(self, path: str) -> None:
        """Render and write workbook to a file path.

        Raises:
            OSError: If path is not writable.
        """
        wb = self.render()
        wb.save(path)

    def metadata(self) -> dict[str, Any]:
        """Return sheet inventory and aggregate stats."""
        total_cells = sum(s.cell_count for s in self._sheets.values())
        return {
            "company_name": self.company_name,
            "ticker": self.ticker,
            "currency": self.currency,
            "report_date": self.report_date,
            "sheet_count": len(self._sheets),
            "sheets": [
                {
                    "name": n,
                    "sheet_type": s.sheet_type,
                    "row_count": s.row_count,
                    "col_count": s.col_count,
                    "cell_count": s.cell_count,
                }
                for n, s in self._sheets.items()
            ],
            "total_cells": total_cells,
        }


# ---------------------------------------------------------------------------
# Pre-built template factories
# ---------------------------------------------------------------------------


class EquityResearchInitiation(WorkbookTemplate):
    """Pre-built skeleton for an equity research initiation workbook.

    Declares standard sheets on construction:
      Cover, Assumptions, DCF, Sensitivity, Scenarios,
      Comparables, SOTP, Cost of Capital, Decision, Citations.

    Full template (formulas, named ranges, conditional formatting) lands in Phase 3.
    """

    _SHEETS: ClassVar[list[tuple[str, SheetType]]] = [
        ("Assumptions", "assumptions"),
        ("DCF", "table"),
        ("Sensitivity", "table"),
        ("Scenarios", "table"),
        ("Comparables", "table"),
        ("SOTP", "table"),
        ("Cost of Capital", "table"),
        ("Decision", "table"),
        ("Citations", "narrative"),
    ]

    def __init__(
        self,
        company_name: str,
        ticker: str,
        currency: str = "USD",
        report_date: str | date | None = None,
    ) -> None:
        super().__init__(company_name, ticker, currency, report_date)
        for name, sheet_type in self._SHEETS:
            self.add_sheet(name, sheet_type=sheet_type)


class EarningsUpdate(WorkbookTemplate):
    """Pre-built skeleton for an earnings update workbook.

    Declares standard sheets:
      Cover, Financials, EPS Walk, Guidance, Citations.

    Full template lands in Phase 3.
    """

    _SHEETS: ClassVar[list[tuple[str, SheetType]]] = [
        ("Financials", "table"),
        ("EPS Walk", "table"),
        ("Guidance", "table"),
        ("Citations", "narrative"),
    ]

    def __init__(
        self,
        company_name: str,
        ticker: str,
        currency: str = "USD",
        report_date: str | date | None = None,
    ) -> None:
        super().__init__(company_name, ticker, currency, report_date)
        for name, sheet_type in self._SHEETS:
            self.add_sheet(name, sheet_type=sheet_type)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _parse_markdown_table_into(spec: SheetSpec, markdown: str) -> None:
    """Parse a pipe-delimited markdown table and load into SheetSpec.

    Handles:
      | Header A | Header B |
      |---|---|
      | row1a | row1b |
    """
    lines = [ln.strip() for ln in markdown.strip().splitlines() if ln.strip()]
    rows_parsed: list[list[str]] = []
    for line in lines:
        if line.startswith("|"):
            cells = [c.strip() for c in line.split("|")[1:-1]]
            # Skip separator lines (--- only)
            if all(set(c).issubset({"-", " ", ":"}) for c in cells):
                continue
            rows_parsed.append(cells)

    if not rows_parsed:
        return
    spec.set_headers(rows_parsed[0])
    for row in rows_parsed[1:]:
        spec.add_row(row)
